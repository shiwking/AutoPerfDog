#!/bin/bash
# -*- coding: utf-8 -*-
import os
import sys

rootPath = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))  # 当前文件所在的目录
sys.path.append(rootPath)
print(rootPath)

import grpc
import openpyxl
import pymysql
import subprocess
import traceback
import platform
import threading
from multiprocessing import Process, Queue
from queue import Empty
import PerfDogService_Linux.demo.python.perfdog_pb2_grpc as perfdog_pb2_grpc
import PerfDogService_Linux.demo.python.perfdog_pb2 as perfdog_pb2
from airtest.core.api import *
from airtest.core.android.adb import ADB
from tools.atx_api.atx2_api import get_device_source_info
from tools.sofunny.autoBattlePerf import AutoBattlePerf
from poco.drivers.unity3d import UnityPoco
from tools.sofunny.command import Command
from setup import getpropath
from MyConfigParser import MyConfigParser, getvaule, setvaule

system = platform.system()

if system == "Windows":
    PerfDogService_exe_path = os.path.join(getpropath()[0], "PerfDogService", "PerfDogService.exe")
else:
    PerfDogService_exe_path = os.path.join(getpropath()[0], "PerfDogService_Linux", "PerfDogService")

q = Queue(1)
adb = ADB().adb_path
configPath = os.path.join(getpropath()[0], "cfg", "config.ini")
build_type = getvaule(configPath, "config", "build_type")
package = getvaule(configPath, "active_config", build_type)
report_output_path = os.path.join(getpropath()[1], "perfdog_service_output",
                                  "AutoPerformance_Report_" + str(time.time()))

personal_token = "866cdeec9743460abd3f1d7e2075300eddfeb964d481e387682361754f582b7c"


def get_performance_data(pack_name, q):
    """
    启动PerfDogService
    测试数据采集
    """
    try:
        # 在代码里启动PerfDogService或手动启动PerfDogService
        print("0.启动PerfDogService")
        # 填入PerfDogService的路径
        perfDogService = subprocess.Popen(PerfDogService_exe_path)
        # 等待PerfDogService启动完毕
        time.sleep(5)
        print("1.通过ip和端口连接到PerfDog Service")
        options = [('grpc.max_receive_message_length', 100 * 1024 * 1024)]
        channel = grpc.insecure_channel('127.0.0.1:23456', options=options)
        print("2.新建一个stub,通过这个stub对象可以调用所有服务器提供的接口")
        stub = perfdog_pb2_grpc.PerfDogServiceStub(channel)
        print("3.通过令牌登录，令牌可以在官网申请")
        userInfo = stub.loginWithToken(
            perfdog_pb2.Token(token=personal_token))
        print("UserInfo:\n", userInfo)
        q.put("start_con")
        print("4.启动设备监听器监听设备,每当设备插入和移除时会收到一个DeviceEvent")
        deviceEventIterator = stub.startDeviceMonitor(perfdog_pb2.Empty())
        print(deviceEventIterator)
        for deviceEvent in deviceEventIterator:
            # 从DeviceEvent中获取到device对象，device对象会在后面的接口中用到
            device = deviceEvent.device
            if deviceEvent.eventType == perfdog_pb2.ADD:
                print("设备[%s:%s]插入\n" % (device.uid, perfdog_pb2.DEVICE_CONTYPE.Name(device.conType)))
                # 每台手机会返回两个conType不同的设备对象(USB的和WIFI的),如果是测有线，取其中的USB对象
                if device.conType == perfdog_pb2.WIFI or device.conType == perfdog_pb2.USB:
                    print("5.初始化设备[%s:%s]\n" % (device.uid, perfdog_pb2.DEVICE_CONTYPE.Name(device.conType)))
                    stub.initDevice(device)
                    print("6.获取app列表")
                    q.put("get_app_list")
                    appList = stub.getAppList(device)
                    apps = appList.app
                    app_index = 0
                    for app in apps:
                        print('%s: %s->%s' % (app_index, app.label, app.packageName))
                        if app.packageName == pack_name:
                            app_select = app_index
                            break
                        else:
                            app_index += 1
                            app_select = None
                    if app_select is None: app_select = int(input("未安装输入APP，请选择要测试App: "))

                    print("7.获取设备的详细信息")

                    deviceInfo = stub.getDeviceInfo(device)
                    print("8.开启性能数据项")
                    stub.enablePerfDataType(
                        perfdog_pb2.EnablePerfDataTypeReq(device=device, type=perfdog_pb2.NETWORK_USAGE))
                    print("9.开始收集[%s:%s]的性能数据\n" % (app.label, app.packageName))
                    print(stub.startTestApp(perfdog_pb2.StartTestAppReq(device=device, app=app)))

                    req = perfdog_pb2.OpenPerfDataStreamReq(device=device)
                    perfDataIterator = stub.openPerfDataStream(req)

                    # def perf_data_process():
                    #     for perfData in perfDataIterator:
                    #         print(perfData)
                    #         sys.stdout.write(perfData)
                    #
                    # threading.Thread(target=perf_data_process).start()

                    time.sleep(10)

                    war_end_sign = q.get(timeout=1800)
                    if war_end_sign == "report_end":
                        pass

                    print("10.设置label")
                    try:
                        stub.setLabel(perfdog_pb2.SetLabelReq(device=device, label="AutoPerfDogTest"))
                    except:
                        while True:
                            try:
                                stub.loginWithToken(
                                    perfdog_pb2.Token(token=personal_token))
                                stub.setLabel(perfdog_pb2.SetLabelReq(device=device, label="AutoPerfDogTest"))
                                break
                            except:
                                continue
                    time.sleep(3)
                    print("11.添加批注")
                    try:
                        stub.addNote(perfdog_pb2.AddNoteReq(device=device, time=5000, note="AutoPerfDogTest"))
                    except:
                        while True:
                            try:
                                stub.loginWithToken(
                                    perfdog_pb2.Token(token=personal_token))
                                stub.addNote(perfdog_pb2.AddNoteReq(device=device, time=5000, note="AutoPerfDogTest"))
                                break
                            except:
                                continue

                    print("12.上传和导出所有数据")
                    try:
                        saveResult = stub.saveData(perfdog_pb2.SaveDataReq(
                            device=device,
                            caseName="case1",  # web上case和excel的名字
                            uploadToServer=True,  # 上传到perfdog服务器
                            exportToFile=True,  # 保存到本地
                            outputDirectory=report_output_path,
                            dataExportFormat=perfdog_pb2.EXPORT_TO_JSON
                        ))
                    except:
                        while True:
                            try:
                                stub.loginWithToken(
                                    perfdog_pb2.Token(token=personal_token))
                                saveResult = stub.saveData(perfdog_pb2.SaveDataReq(
                                    device=device,
                                    caseName="case1",  # web上case和excel的名字
                                    uploadToServer=True,  # 上传到perfdog服务器
                                    exportToFile=True,  # 保存到本地
                                    outputDirectory=report_output_path,
                                    dataExportFormat=perfdog_pb2.EXPORT_TO_JSON
                                ))
                                break
                            except:
                                continue

                    print("保存结果:\n", saveResult)
                    print("12.上传和导出第所有数据excel")
                    try:
                        stub.saveData(perfdog_pb2.SaveDataReq(
                            device=device,
                            caseName="case2",  # web上case和excel的名字
                            uploadToServer=True,  # 上传到perfdog服务器
                            exportToFile=True,  # 保存到本地
                            outputDirectory=report_output_path,
                            dataExportFormat=perfdog_pb2.EXPORT_TO_EXCEL
                        ))
                    except:
                        while True:
                            try:
                                stub.loginWithToken(
                                    perfdog_pb2.Token(token=personal_token))
                                stub.saveData(perfdog_pb2.SaveDataReq(
                                    device=device,
                                    caseName="case2",  # web上case和excel的名字
                                    uploadToServer=True,  # 上传到perfdog服务器
                                    exportToFile=True,  # 保存到本地
                                    outputDirectory=report_output_path,
                                    dataExportFormat=perfdog_pb2.EXPORT_TO_EXCEL
                                ))
                                break
                            except:
                                continue

                    print("保存结果:\n", saveResult)
                    print("13.测试结果写入数据库")
                    try:
                        write_performance_data_to_sql(report_output_path)
                    except:
                        pass
                    print("13.停止测试")
                    stub.stopTest(perfdog_pb2.StopTestReq(device=device))
                    print("over")
                    break
            elif deviceEvent.eventType == perfdog_pb2.REMOVE:
                sys.stdout.write("设备[%s:%s]移除\n" % (device.uid, perfdog_pb2.DEVICE_CONTYPE.Name(device.conType)))
    except Exception as e:
        print(e)
        traceback.print_exc()


def write_performance_data_to_sql(report_output_path):
    sys.stdout.write(report_output_path)
    wb = openpyxl.load_workbook(report_output_path, data_only=True)
    sheet = wb["all"]
    pack_name = package.split("_")[1].replace("P", "").replace("S", "")
    device_name = sheet.cell(5, 1).value
    FPS_AVG = sheet.cell(9, 1).value
    Jank = sheet.cell(9, 6).value
    BigJank = sheet.cell(9, 7).value
    Memory_AVG = sheet.cell(9, 10).value
    Memory_Peak = sheet.cell(9, 11).value
    AppCPU = sheet.cell(9, 13).value
    wb.close()

    if FPS_AVG < 40 or Jank > 62 or BigJank > 37 or Memory_AVG > 1054 or Memory_Peak > 1099:
        result = "Fail"
    else:
        result = "Pass"

    sys.stdout.write(pack_name, device_name, FPS_AVG, Jank, BigJank, Memory_AVG, Memory_Peak, AppCPU, result)

    db = pymysql.connect(host="10.30.20.29", port=3306, user="root", password="@Root123", database="reprot")
    cursor = db.cursor()

    select_sql = "select id FROM perfAutotest ORDER BY id DESC LIMIT 0,1 "

    try:
        # 执行SQL语句
        cursor.execute(select_sql)
        # 获取所有记录列表
        results = cursor.fetchall()
        sys.stdout.write(results)
        for i in results:
            new_id = i[0]
    except Exception as e:
        sys.stdout.write(e)
        sys.stdout.write("Error: unable to fetch data")

    sql = "INSERT INTO perfAutotest (testid,testdevice, testpakname, fps, jank,bigjank,memory,peakmemory, cpu,testresult,isDelete) " \
          "VALUES ('%s', '%s',  '%s',  '%s','%s', '%s', '%s','%s','%s','%s','%s')" % (
              int(new_id) + 1, device_name, pack_name, FPS_AVG, Jank, BigJank, Memory_AVG, Memory_Peak, AppCPU,
              result,
              0)
    try:
        # 执行sql语句
        cursor.execute(sql)
        # 执行sql语句提交
        db.commit()
    except Exception as e:
        # 发生错误时回滚
        sys.stdout.write("Error")
        sys.stdout.write(e)
        db.rollback()
    db.close()


def start_war_report():
    poco = UnityPoco()
    poco.agent.command = Command(poco.agent.c)
    autoRunMap = AutoBattlePerf(poco.agent.command)
    autoRunMap.initModel()
    autoRunMap.playReport(warID="27e1407c4d796607cec958141d4fa88a_1", playID=385015877, wartype=1)


def cheak_is_end(q):
    poco = UnityPoco()
    poco.agent.command = Command(poco.agent.c)
    autoRunMap = AutoBattlePerf(poco.agent.command)
    autoRunMap.initModel()
    time.sleep(15)
    while True:
        result = autoRunMap.isEndReport()
        if result["status"] == 2:
            q.put("report_end")
            break
        else:
            time.sleep(3)
            continue


if __name__ == '__main__':
    values = sys.argv
    dev = values[1]
    dev_connect_url = get_device_source_info(dev)["device"]["source"]["remoteConnectAddress"]
    print(dev_connect_url)

    connect_device("Android://127.0.0.1:5037/" + dev_connect_url)
    stop_app(package=package)
    start_app(package=package)
    os.system(adb + " devices")
    time.sleep(40)
    start_war_report()
    time.sleep(20)

    get_performance_data_process = Process(target=get_performance_data, args=(package, q))
    get_performance_data_process.start()

    if q.get() == "start_con":
        pass

    while True:
        time.sleep(5)
        os.system(adb + " connect " + dev_connect_url)
        try:
            if q.get_nowait() == "get_app_list":
                print("device reconnect success!!")
                break
        except Empty as e:
            print("q is Empty,try again")
            pass

    cheak_is_end(q)

    get_performance_data_process.join()

    connect_device("Android://127.0.0.1:5037/" + dev_connect_url)
    stop_app(package=package)
